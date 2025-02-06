import sys
from pathlib import Path
from datetime import datetime
import argparse
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Side, Border, Alignment, Font
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from ups_tool import util as fn_util

###########################################
# 분포
#   input
#       data schema xlsx: 번호, 컬럼명, 컬럼설명
#       dist folder: 컬럼별 분포 csv
#   output
#       기초자료 분포 xlsx

shorten_lines = 21      # 21행 이하인 경우 전부. 초과인 경우 20행(10+빈줄+10)
schema_columns = ['no', 'name', 'desc']
info_columns = ['컬럼번호', '컬럼명', '컬럼 설명']
dist_columns = ['SEQ', '항목값', '빈도', '비율']
style_background = PatternFill(start_color='00eeeeee', fill_type='solid')
style_line = Side(border_style='thin', color='00000000')
style_border = Border(top=style_line, bottom=style_line, left=style_line, right=style_line)
xls_columns = info_columns + dist_columns + ['비고']
xls_info_idx = [0, 1, 2]    # 컬럼번호, 컬럼명, 컬럼 설명
xls_data_idx = list(set(x for x in range(len(xls_columns))) - set(xls_info_idx)) # [3, 4, 5, 6] # SEQ, 항목값, 빈도, 비율, 비고
start_column = 1
start_row = 5

kinds = {
    'old': {
        'title': '07. 원본 데이터 분포(범주형)',
        'start_column': start_column,
        'end_column': start_column + len(xls_columns) - 1,
        'style': {},
    },
    'new': {
        'title': '07. 가명처리 후 데이터 분포(범주형)',
        'start_column': start_column + len(xls_columns) + 1,
        'end_column': start_column + len(xls_columns) * 2,
        'style': {},
    },
}
for kind, kv in kinds.items():
    for x in range(len(xls_columns)):
        kv['style'][kv['start_column']+0] = {'width': 12.3, 'alignment': Alignment(horizontal='center', vertical='top')}
        kv['style'][kv['start_column']+1] = {'width': 16.3, 'alignment': Alignment(horizontal='center', vertical='top')}
        kv['style'][kv['start_column']+2] = {'width': 27.5, 'alignment': Alignment(horizontal='center', vertical='top')}
        kv['style'][kv['start_column']+3] = {'width': 12.5, 'alignment': Alignment(horizontal='center', vertical='top')}
        kv['style'][kv['start_column']+4] = {'width': 27.5, 'alignment': Alignment(horizontal='center', vertical='top')}
        kv['style'][kv['start_column']+5] = {'width':  9.5, 'number_format': '#,##0', 'alignment': Alignment(horizontal='right', vertical='top')}
        kv['style'][kv['start_column']+6] = {'width':  9.0, 'number_format': '0.0000%', 'alignment': Alignment(horizontal='left', vertical='top')}
        kv['style'][kv['start_column']+7] = {'width':  5.0, 'alignment': Alignment(horizontal='center', vertical='top')}

###########################################
get_xls_index = lambda c, r: f'{get_column_letter(c)}{r}'

wb = Workbook()
ws = wb.active
ws.title = '12.분포(범주형)'

###########################################
def main(args):
    # 분포 파일 목록
    make_file_map(args)
    print('\n'.join([f'{k}: {v}' for k, v in vars(args).items()]))

    # 컬럼 정보
    df_schema = read_schema(args)
    df_schema = df_schema.sort_values(schema_columns[0])
    df_schema['name'] = df_schema['name'].apply(lambda x: f'{x.strip()}')
    # print(df_schema)

    # 작성방법 때문에 몹시 방해됨 => 행 고정 해제
    # ws.freeze_panes = get_xls_index(1, start_row)    # 행만 고정

    # 헤더
    for _, kv in kinds.items():
        ws_header(args, kv)

    current_row = start_row

    # 분포
    for _, schema in df_schema.iterrows():
        # load 분포 csv
        # filename = f'{args.catalog}.{schema.iloc[1].lower()}'
        filename = f'{args.catalog}.{schema.iloc[1]}'
        print(f'filename: {filename}')
        get_dataframe(kinds['old'], filename, args)
        get_dataframe(kinds['new'], filename, args)
        if (kinds['old']['df'] is None) and (kinds['new']['df'] is None):
            continue

        print(f'{schema.iloc[0]:>4} {schema.iloc[1]}')

        # 제목/헤더
        for _, kv in kinds.items():
            if kv['df'] is None: kv['df'] = pd.DataFrame(columns=dist_columns)
            # 컬럼정보: 번호, 컬럼명, 컬럼설명
            ws_column_info(start_row=current_row, schema=schema, kv=kv)

        # 분포: 값, 개수, 비율
        old_end_row = ws_distribution(start_row=current_row, kv=kinds['old'])
        new_end_row = ws_distribution(start_row=current_row, kv=kinds['new'])
        current_row = max(old_end_row, new_end_row)

    ws_style()

    Path(args.out_xls).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out_xls)
    print(args.out_xls)
    return args.out_xls

###########################################
def ws_distribution(start_row, kv):
    start_column = kv['start_column']
    current_row = start_row
    for idx, row in kv['df'].iterrows():
        # print(row)
        # seq, 항목값, 빈도, 비율
        for x in range(row.size):
            ws_cell(current_row, start_column+xls_data_idx[x], row.iloc[x]) # 3. SEQ
        current_row += 1
    ws_border(min_row=start_row, max_row=current_row-1, min_col=start_column, max_col=start_column+len(xls_columns)-1)
    return current_row

###########################################
def ws_column_info(start_row, schema, kv):
    if len(kv['df']) == 0:
        return

    start_column = kv['start_column']
    merge_size = len(kv['df'])
    for ii, x in enumerate(xls_info_idx):
        if merge_size > 1:
            ws.merge_cells(f'{get_xls_index(start_column+x, start_row)}:{get_xls_index(start_column+x, start_row+merge_size-1)}')
        ws_cell(start_row, start_column+x, schema.iloc[ii])

###########################################
def ws_header(args, kv):
    ss = '연속형 데이터의 분포를 기재'
    ss += '\n데이터별 No.는 하나의 컬럼에 각기다른 값에 대해 1부터 부여'
    ss += '\nNo., 컬럼명, 컬럼설명은 활용데이터 요구 수준표의 컬럼명과 컬럼설명, No를 그대로 사용할 것'
    ss += '\n개수는 각 값을 가지고 있는 사람의 총합을 기재'
    ss += '\n비율은 그 컬럼의 (특정 값의 개수/그 컬럼의 모든 값의 개수의 합)을 %로 표기'

    for _, kv in kinds.items():
        # 1
        ws.merge_cells(f'{get_xls_index(kv["start_column"], 1)}:{get_xls_index(kv["end_column"], 1)}')
        ws_cell(1, kv['start_column'], kv['title'])
        # 2
        ws_cell(2, kv['start_column'], '작성방법')
        ws.merge_cells(f'{get_xls_index(kv["start_column"]+1, 2)}:{get_xls_index(kv["end_column"], 2)}')
        ws_cell(2, kv['start_column']+1, ss)
        # 3
        ws.merge_cells(f'{get_xls_index(kv["start_column"], 3)}:{get_xls_index(kv["start_column"]+2, 3)}')
        ws_cell(3, kv['start_column'], '테이블명')
        ws.merge_cells(f'{get_xls_index(kv["start_column"]+3, 3)}:{get_xls_index(kv["end_column"], 3)}')
        ws_cell(3, kv['start_column']+3, args.catalog)
        # 4
        for ii, v in enumerate(xls_columns):
            ws_cell(4, kv['start_column']+ii, v, isbg=True)
        ws_border(min_row=1, max_row=4, min_col=kv['start_column'], max_col=kv['end_column'])

###########################################
def ws_cell(r, c, v, isbg=False):
    wc = ws[get_xls_index(c, r)]
    wc.value = v
    if isbg: wc.fill = style_background

###########################################
def ws_border(min_row, max_row, min_col, max_col):
    for r in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for c in r:
            c.border = style_border

###########################################
def ws_style():
    # 열 너비
    for _, kv in kinds.items():
        for idx, w in kv['style'].items():
            ws.column_dimensions[get_column_letter(idx)].width = w['width']

    # 행 높이
    ws.row_dimensions[2].height = 71    # 작성방법

    # 제목/헤더
    for _, kv in kinds.items():
        for r in ws.iter_rows(min_row=1, max_row=start_row-1, min_col=kv['start_column'], max_col=kv['end_column']):
            for c in r:
                c.alignment = Alignment(horizontal='center', vertical='center')
        # 제목
        wc = ws[get_xls_index(kv['start_column'], 1)]
        wc.font = Font(size=16, bold=True)
        # 작성방법 내용
        wc = ws[get_xls_index(kv['start_column']+1, 2)]
        wc.font = Font(size=9)
        wc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 분포
    for _, kv in kinds.items():
        for idx, st in kv['style'].items():
            for r in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=idx, max_col=idx):
                for c in r:
                    c.alignment = st['alignment']
                    if 'number_format' in st:
                        c.number_format = st['number_format']

    # merged cell border 없어지는 문제 해결
    for cs in ws.merged_cells.ranges:
        for r in ws.iter_rows(min_row=cs.min_row, max_row=cs.max_row, min_col=cs.min_col, max_col=cs.max_col):
            for c in r:
                c.border = style_border

###########################################
def get_dataframe(kv, name, args):
    if name not in kv['files']:
        # print(f'{name} not in {kv["files"]}')
        kv['df'] = None
        return

    # print('get_dataframe:', kv['files'][name])

    # 항목값, 빈도
    # df = pd.read_csv(kv['files'][name], header=0, encoding=args.encoding, dtype=str)
    df = pd.read_csv(kv['files'][name], header=0, encoding=args.encoding)
    # 빈도 => 정수
    df.iloc[:,1] = df.iloc[:,1].astype('int64')
    # 비율 컬럼 추가
    sumv = df.iloc[:,1].sum(axis=0)
    df['ratio'] = df.iloc[:,1].apply(lambda x: x / sumv).astype('float64')
    # Sort & SEQ 컬럼 추가
    # 정렬: value(0) or count(1)
    sort_columns = [1,0] if args.sort_count else [0,1]
    df = df.sort_values(by=[df.columns[cidx] for cidx in sort_columns])
    df.index = np.arange(1, len(df)+1)
    df['seq'] = df.index
    # 컬럼 순서 변경: SEQ, 항목값, 빈도, 비율
    df = df[df.columns[-1:].to_list() + df.columns[:-1].to_list()]
    # # 컬럼명 변경
    # df.columns = dist_columns

    kv['df'] = make_shorten(df) if (args.shorten and (len(df) > shorten_lines)) or (len(df) > args.max_rows) else df

###########################################
def make_shorten(df):
    return pd.concat([df.head(shorten_lines//2),
                    pd.DataFrame([['...' for x in df.columns]], columns=df.columns),
                    df.tail(shorten_lines//2)], ignore_index=True)

###########################################
def make_file_map(args):
    # kinds['old']['files'] = {x.stem.lower(): x for x in Path(args.dist_old).glob('*') if x.suffix.lower()=='.csv'}
    # kinds['new']['files'] = {}
    # if 'dist_new' in args:
    #     kinds['new']['files'] = {x.stem.lower(): x for x in Path(args.dist_new).glob('*') if x.suffix.lower()=='.csv'}
    kinds['old']['files'] = {x.stem: x for x in Path(args.dist_old).glob('*') if x.suffix.lower()=='.csv'}
    kinds['new']['files'] = {}
    if 'dist_new' in args:
        kinds['new']['files'] = {x.stem: x for x in Path(args.dist_new).glob('*') if x.suffix.lower()=='.csv'}

# ###########################################
# # openpyxl 사용 시 불필요
# def column_number(xls_index):
#     # 26진법: A = 1, Z = 26
#     # BC = B*26 + C = 2*26 + 3 = 55
#     _to_number = lambda x: ord(x.upper()) - ord('A') + 1

#     idx = 0
#     xs = list(xls_index)
#     xs.reverse()
#     for ii, x in enumerate(xs):
#         idx += 26**ii * _to_number(x)
#     return idx - 1

###########################################
def read_schema(args):
    print(args)

    # 시트명이 31자 제한이 왜 있는건지...
    with pd.ExcelFile(args.schema_file) as xx:
        #  print(xx.sheet_names)
         sheet_idx = xx.sheet_names.index(args.sheet[:31])

    import warnings
    with warnings.catch_warnings():
        warnings.filterwarnings('ignore', category=UserWarning)
        df = pd.read_excel(
                io=args.schema_file,
                sheet_name=sheet_idx,
                # sheet_name=args.sheet,
                header=args.header - 1,
                index_col=None,
                dtype=str,
                engine='openpyxl',
        )
    idxs = [column_index_from_string(x)-1 for x in [args.number, args.name, args.desc]]
    df = df.iloc[:,idxs]
    df.columns = schema_columns
    df = df.astype({schema_columns[0]:'int32'})
    return df

###########################################
if __name__ == "__main__":
    def _min_max_rows(x):
        x = int(x)
        if x < shorten_lines:
            raise argparse.ArgumentTypeError(f'max_rows는 {shorten_lines}보다 커야 합니다')
        return x

    parser = argparse.ArgumentParser(formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument('--catalog', required=True, help='테이블이름')
    parser.add_argument('--encoding', required=False, default='euc-kr', help='분포 csv 인코딩 (기본값: %(default)s)')
    parser.add_argument('--dist_old', required=True, action=fn_util.PathAction, help='원본 데이터 컬럼별 분포 csv를 포함한 디렉토리')
    parser.add_argument('--dist_new', required=True, action=fn_util.PathAction, help='처리 후 데이터 컬럼별 분포 csv를 포함한 디렉토리')
    parser.add_argument('--schema_file', required=False, action=fn_util.PathAction, help='컬럼 정보를 포함한 액셀 파일')
    parser.add_argument('--sheet', required=False, default='06. 원본 데이터 명세(상세) 및 예시', help='액셀 시트명 (기본값: %(default)s)')
    parser.add_argument('--header', required=False, type=int, default=3, help='시트 내 헤더 행 번호 (기본값: %(default)s)')
    parser.add_argument('--number', required=False, default='A', help='"컬럼번호" 열 인덱스 (기본값: %(default)s)')
    parser.add_argument('--name', required=False, default='B', help='"컬럼이름" 열 인덱스 (기본값: %(default)s)')
    parser.add_argument('--desc', required=False, default='C', help='"컬럼설명" 열 인덱스 (기본값: %(default)s)')
    parser.add_argument('--out_xls', required=False, default=Path(f'./기초자료.분포.{fn_util.get_file_postfix()}.xlsx'),
                        action=fn_util.PathAction, help='결과 액셀 파일 (기본값: 기초자료.분포.<실행시간>.xlsx)')
    parser.add_argument('--shorten', required=False, default=False, action=argparse.BooleanOptionalAction, help=f'데이터 행 축약 여부\n{shorten_lines}행보다 많으면 축약 (기본값: %(default)s)')
    parser.add_argument('--max_rows', default=1000, type=_min_max_rows, help=f'최대 행 수\n최대 행 수보다 많으면 축약\n최소 {shorten_lines}보다 커야 함')
    parser.add_argument('--sort_count', default=True, action=argparse.BooleanOptionalAction, help='정렬 기준: count or value (기본값: --sort_count)')
    args = parser.parse_args()
    main(args)
