from openpyxl import Workbook
from openpyxl.styles import PatternFill, Side, Border, Alignment, Font
from openpyxl.utils.cell import get_column_letter, column_index_from_string

class XLS:
    def __init__(self, font_size=9) -> None:
        self.wb = Workbook()
        self.ws = self.wb.active
        self.font_size = font_size

    ###########################################
    cell_name = lambda self, col_no, row_no: f'{get_column_letter(col_no)}{row_no}'
    col_no = lambda self, col_char: column_index_from_string(col_char)
    col_letter = lambda self, col_no: get_column_letter(col_no)

    ###########################################
    def get_cell(self, col_no, row_no):
        return self.ws.cell(row=row_no, column=col_no)

    ###########################################
    def set_cell(self, col_no, row_no, value, font_size=None, bg_color=None):
        wc = self.ws.cell(row=row_no, column=col_no, value=value)
        wc.font = Font(size=font_size if font_size else self.font_size)
        if bg_color:
            wc.fill = PatternFill(start_color=f'00{bg_color}', fill_type='solid')
        return wc

    ###########################################
    def merge_cells(self, start_col_no, start_row_no, end_col_no, end_row_no):
        self.ws.merge_cells(start_row=start_row_no, start_column=start_col_no,
                         end_row=end_row_no, end_column=end_col_no)

    def merge_cells(self, start_cell, end_cell):
        self.ws.merge_cells(range_string=f'{start_cell}:{end_cell}')

    ###########################################
    def alignment(self, start_col_no, start_row_no, end_col_no, end_row_no, align='center', valign='center'):
        for r in self.ws.iter_rows(min_row=start_row_no, max_row=end_row_no, min_col=start_col_no, max_col=end_col_no):
            for c in r:
                c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=True)

    ###########################################
    def border(self, start_col_no, start_row_no, end_col_no, end_row_no,
               border_style='thin', border_color='000000'):
        style_line = Side(border_style=border_style, color=f'00{border_color}')
        style_border = Border(top=style_line, bottom=style_line, left=style_line, right=style_line)
        for r in self.ws.iter_rows(min_row=start_row_no, max_row=end_row_no, min_col=start_col_no, max_col=end_col_no):
            for c in r:
                c.border = style_border

    ###########################################
    def col_width(self, col_no, width):
        self.ws.column_dimensions[get_column_letter(col_no)].width = width

    ###########################################
    def row_height(self, row_no, height):
        self.ws.row_dimensions[row_no].height = height

    ###########################################
    def set_sheet_name(self, name):
        self.ws.title = name

    ###########################################
    def create_sheet(self, name):
        self.ws = self.wb.create_sheet()
        self.set_sheet_name(name)

    ###########################################
    def active_sheet(self, name):
        self.ws = self.wb[name]

    ###########################################
    def save(self, path):
        self.wb.save(path)
